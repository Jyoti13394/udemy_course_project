import inspect
import logging

from openpyxl import load_workbook


class Utilities:

    def data_from_excel_to_dict(path):
        data_dict = {}
        wb = load_workbook(path)
        sheet = wb.active
        rows = sheet.max_row
        for i in range(1, rows+1):
            #cell_obj = sheet.cell(row=i, column=1)
            data_dict[sheet.cell(row=i, column=1).value] = sheet.cell(row=i, column=2).value
        return data_dict

    def getLogger(self):
        loggername = inspect.stack()[1][3]
        logger = logging.getLogger(loggername)
        filehandler = logging.FileHandler('logfile.log')
        formatter = logging.Formatter("%(asctime)s : %(name)s : %(levelname)s : %(message)s", datefmt='%m/%d/%Y %I:%M:%S %p')
        filehandler.setFormatter(formatter)
        logger.addHandler(filehandler)
        logger.setLevel(logging.DEBUG)
        return logger
